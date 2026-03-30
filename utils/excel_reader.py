from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from core.exceptions import DataFileFormatError


TRUE_VALUES = {"1", "true", "y", "yes", "on"}
DEFAULT_REQUIRED_COLUMNS = {
    "execute",
    "data_id",
    "test_title",
    "expected_result",
    "checkpoints",
}
JSON_COLUMNS = {"steps_data"}


def _normalize_value(value: Any) -> Any:
    if value is None:
        return ""
    return value


def _is_enabled(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in TRUE_VALUES


class ExcelReader:
    def __init__(self, file_path: str | Path, required_columns: set[str] | None = None) -> None:
        self.file_path = Path(file_path)
        self.required_columns = required_columns or DEFAULT_REQUIRED_COLUMNS

    @staticmethod
    def _normalize_headers(values: list[Any]) -> list[str]:
        return [str(value).strip() if value is not None else "" for value in values]

    @staticmethod
    def _convert_cell(header: str, value: Any) -> Any:
        normalized_value = _normalize_value(value)
        if header in JSON_COLUMNS and isinstance(normalized_value, str):
            text = normalized_value.strip()
            if text.startswith("{") or text.startswith("["):
                try:
                    return json.loads(text)
                except json.JSONDecodeError as error:
                    raise DataFileFormatError(
                        f"Column '{header}' in '{text}' is not valid JSON."
                    ) from error
        return normalized_value

    def read_rows(self, sheet_name: str = "data") -> list[dict[str, Any]]:
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        workbook = load_workbook(filename=self.file_path, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise DataFileFormatError(f"Sheet '{sheet_name}' not found in {self.file_path.name}")

            worksheet = workbook[sheet_name]
            values = list(worksheet.iter_rows(values_only=True))
        finally:
            workbook.close()

        if not values:
            return []

        headers = self._normalize_headers(list(values[0]))
        missing_columns = sorted(column for column in self.required_columns if column not in headers)
        if missing_columns:
            raise DataFileFormatError(
                f"Excel file '{self.file_path.name}' is missing required columns in sheet '{sheet_name}': "
                f"{', '.join(missing_columns)}"
            )

        rows: list[dict[str, Any]] = []
        for raw_row in values[1:]:
            row = {
                headers[index]: self._convert_cell(headers[index], value)
                for index, value in enumerate(raw_row)
                if index < len(headers) and headers[index]
            }
            if any(str(value).strip() for value in row.values()):
                rows.append(row)
        return rows

    def get_executable_rows(self, sheet_name: str = "data") -> list[dict[str, Any]]:
        rows = self.read_rows(sheet_name=sheet_name)
        return [row for row in rows if _is_enabled(row.get("execute"))]

